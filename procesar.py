"""
Procesador automático de contenedores
- Lee fotos nuevas desde OneDrive
- Filtra con Claude (2 etapas)
- Actualiza el Excel en OneDrive
- Registra qué fotos ya fueron procesadas (processed.json)
"""

import google.generativeai as genai
import io
import json
import os
import sys
import requests
from datetime import datetime
from PIL import Image

# -------------------------------------------------------
# CONFIGURACIÓN — variables de entorno (GitHub Secrets)
# -------------------------------------------------------
GEMINI_KEY         = os.environ["GEMINI_KEY"]
ONEDRIVE_TOKEN_URL = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
CLIENT_ID          = os.environ["MS_CLIENT_ID"]
CLIENT_SECRET      = os.environ["MS_CLIENT_SECRET"]
TENANT_ID          = os.environ["MS_TENANT_ID"]

CARPETA_FOTOS      = os.environ.get("CARPETA_FOTOS", "Contenedores/Fotos")
EXCEL_PATH         = os.environ.get("EXCEL_PATH", "Contenedores/reporte.xlsx")
PROCESSED_FILE     = "processed.json"

# Configurar Gemini
genai.configure(api_key=GEMINI_KEY)
model = genai.GenerativeModel('gemini-2.5-flash')


# -------------------------------------------------------
# AUTENTICACIÓN MICROSOFT GRAPH
# -------------------------------------------------------

def obtener_token() -> str:
    """Obtiene access token de Microsoft Graph (app-only)"""
    url  = ONEDRIVE_TOKEN_URL.format(tenant=TENANT_ID)
    data = {
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
    }
    resp = requests.post(url, data=data)
    resp.raise_for_status()
    return resp.json()["access_token"]


def headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


# -------------------------------------------------------
# OPERACIONES ONEDRIVE
# -------------------------------------------------------

def listar_fotos(token: str) -> list[dict]:
    """Lista archivos de imagen en la carpeta de OneDrive"""
    url  = f"https://graph.microsoft.com/v1.0/me/drive/root:/{CARPETA_FOTOS}:/children"
    resp = requests.get(url, headers=headers(token))
    resp.raise_for_status()
    items = resp.json().get("value", [])

    extensiones = {".jpg", ".jpeg", ".png", ".webp"}
    return [
        {"id": f["id"], "nombre": f["name"], "url": f["@microsoft.graph.downloadUrl"]}
        for f in items
        if any(f["name"].lower().endswith(ext) for ext in extensiones)
    ]


def descargar_foto(url: str) -> bytes:
    """Descarga el contenido binario de una foto"""
    resp = requests.get(url)
    resp.raise_for_status()
    return resp.content


def descargar_excel(token: str) -> bytes:
    """Descarga el Excel actual desde OneDrive"""
    url  = f"https://graph.microsoft.com/v1.0/me/drive/root:/{EXCEL_PATH}:/content"
    resp = requests.get(url, headers=headers(token))
    resp.raise_for_status()
    return resp.content


def subir_excel(token: str, contenido: bytes):
    """Sube el Excel actualizado a OneDrive (sobreescribe)"""
    url  = f"https://graph.microsoft.com/v1.0/me/drive/root:/{EXCEL_PATH}:/content"
    resp = requests.put(
        url,
        headers={**headers(token), "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        data=contenido,
    )
    resp.raise_for_status()


# -------------------------------------------------------
# REGISTRO DE FOTOS YA PROCESADAS
# -------------------------------------------------------

def cargar_procesadas() -> set:
    """Carga el set de nombres de archivo ya procesados"""
    if os.path.exists(PROCESSED_FILE):
        with open(PROCESSED_FILE) as f:
            return set(json.load(f))
    return set()


def guardar_procesadas(procesadas: set):
    """Guarda el set actualizado"""
    with open(PROCESSED_FILE, "w") as f:
        json.dump(list(procesadas), f)


# -------------------------------------------------------
# UTILIDADES DE IMAGEN Y JSON
# -------------------------------------------------------

def preparar_imagen(image_bytes: bytes) -> Image.Image:
    """Abre los bytes descargados y los convierte en un objeto Image para Gemini, optimizando tamaño."""
    img = Image.open(io.BytesIO(image_bytes))
    if img.mode != "RGB":
        img = img.convert("RGB")
    if max(img.size) > 1024:
        img.thumbnail((1024, 1024), Image.LANCZOS)
    return img

def limpiar_json(texto: str):
    """Limpia la respuesta para obtener solo el JSON válido"""
    try:
        s = texto.strip()
        if "```json" in s:
            s = s.split("```json")[1].split("```")[0]
        elif "```" in s:
            s = s.split("```")[1].split("```")[0]
        return json.loads(s)
    except Exception:
        return None

# -------------------------------------------------------
# PIPELINE IA (2 etapas con Gemini)
# -------------------------------------------------------

def es_puerta_contenedor(img: Image.Image) -> bool:
    prompt = "Is this a photo of a shipping container door showing the container code/number? Answer only YES or NO."
    response = model.generate_content([prompt, img])
    return response.text.strip().upper().startswith("YES")

def extraer_datos_contenedor(img: Image.Image):
    prompt = """Actúa como OCR experto en logística. Extrae en JSON:
- sigla, numero, dv, max_gross_kg, tara_kg
Si no es legible pon null. Solo el JSON, sin explicaciones."""
    response = model.generate_content([prompt, img])
    return limpiar_json(response.text)


def extraer_datos_contenedor(image_b64: str):
    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=200,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_b64}},
                {"type": "text",  "text": """Actúa como OCR experto en logística. Extrae en JSON:
- sigla, numero, dv, max_gross_kg, tara_kg
Si no es legible pon null. Solo el JSON, sin explicaciones."""}
            ],
        }],
    )
    return limpiar_json(response.content[0].text)


# -------------------------------------------------------
# ACTUALIZAR EXCEL
# -------------------------------------------------------

def agregar_filas_excel(excel_bytes: bytes, filas: list[dict]) -> bytes:
    """Agrega filas nuevas al Excel existente y devuelve los bytes actualizados"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    # Si el Excel está vacío, crear encabezados
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        encabezados = ["fecha", "archivo_origen", "status", "sigla", "numero", "dv", "max_gross_kg", "tara_kg"]
        ws.append(encabezados)

    for fila in filas:
        ws.append([
            fila.get("fecha", ""),
            fila.get("archivo_origen", ""),
            fila.get("status", ""),
            fila.get("sigla", ""),
            fila.get("numero", ""),
            fila.get("dv", ""),
            fila.get("max_gross_kg", ""),
            fila.get("tara_kg", ""),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# -------------------------------------------------------
# MAIN
# -------------------------------------------------------

def main():
    print("=== Iniciando procesamiento ===")

    token      = obtener_token()
    procesadas = cargar_procesadas()

    # Listar fotos en OneDrive
    todas_las_fotos = listar_fotos(token)
    fotos_nuevas    = [f for f in todas_las_fotos if f["nombre"] not in procesadas]

    if not fotos_nuevas:
        print("No hay fotos nuevas. Fin.")
        return

    print(f"Fotos nuevas encontradas: {len(fotos_nuevas)}")

    # Descargar Excel actual
    excel_bytes = descargar_excel(token)

    filas_nuevas   = []
    nombres_ok     = set()
    procesadas_cnt = 0
    descartadas    = 0
    errores        = 0

    for foto in fotos_nuevas:
        nombre = foto["nombre"]
        print(f"  Procesando: {nombre}")

        try:
            image_bytes = descargar_foto(foto["url"])
            img  = preparar_imagen(image_bytes)

            # Etapa 1: filtro
            if not es_puerta_contenedor(image_b64):
                descartadas += 1
                filas_nuevas.append({
                    "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "archivo_origen": nombre,
                    "status": "Descartada",
                })
                nombres_ok.add(nombre)
                print(f"    → Descartada")
                continue

            # Etapa 2: OCR
            datos = extraer_datos_contenedor(image_b64)

            if datos:
                procesadas_cnt += 1
                filas_nuevas.append({
                    "fecha":          datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "archivo_origen": nombre,
                    "status":         "OK",
                    **datos,
                })
                print(f"    → OK: {datos.get('sigla','?')}{datos.get('numero','?')}")
            else:
                errores += 1
                filas_nuevas.append({
                    "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "archivo_origen": nombre,
                    "status": "Error OCR",
                })
                print(f"    → Error OCR")

            nombres_ok.add(nombre)

        except Exception as e:
            errores += 1
            print(f"    → Error sistema: {e}")
            filas_nuevas.append({
                "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "archivo_origen": nombre,
                "status": f"Error: {e}",
            })
            nombres_ok.add(nombre)

    # Actualizar Excel y subir
    if filas_nuevas:
        excel_actualizado = agregar_filas_excel(excel_bytes, filas_nuevas)
        subir_excel(token, excel_actualizado)
        print(f"Excel actualizado con {len(filas_nuevas)} filas nuevas.")

    # Guardar registro de procesadas
    procesadas.update(nombres_ok)
    guardar_procesadas(procesadas)

    print(f"\n=== Resumen ===")
    print(f"  Extraídas:   {procesadas_cnt}")
    print(f"  Descartadas: {descartadas}")
    print(f"  Errores:     {errores}")


if __name__ == "__main__":
    main()
